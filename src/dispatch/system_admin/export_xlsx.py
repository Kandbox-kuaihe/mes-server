from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from dispatch.system_admin.menu.models import Menu
from dispatch.system_admin.menu.models_secondary import role_menu_button_table
from dispatch.system_admin.menu_button.models import MenuButton
from dispatch.system_admin.role.models import Role
from dispatch.database_util.service import get_schema_session

def export_menus_to_excel(db_session, file_path: str):
    role_menu_buttons = defaultdict(list)
    # 查询所有菜单数据
    menus = db_session.query(Menu).order_by(Menu.id.asc()).all()
    roles = db_session.query(Role).order_by(Role.id.asc()).all()

    for role in roles:
        role_menu_button_objs = db_session.query(role_menu_button_table).filter(role_menu_button_table.c.role_id == role.id).all()
        for menu_button_id_role_id in role_menu_button_objs:
            button_id, _ = menu_button_id_role_id
            role_menu_buttons[role.id].append(button_id)

    # 将查询结果转换为DataFrame
    data = []
    for menu in menus:
        buttons = db_session.query(MenuButton).filter(MenuButton.menu_id == menu.id).all()
        for button in buttons:
            row = {
                'UI Function': menu.name,
                'Operation': button.name
            }
            for role in roles:
                if button.id in role_menu_buttons[role.id]:
                    row[role.name] = True
                else:
                    row[role.name] = False
            data.append(row)
    
    login_function_row = {
        'UI Function': 'Login',
        'Operation': 'Login',
    }
    for role in roles:
        login_function_row[role.name] = True
    data.insert(0, login_function_row)
    df = pd.DataFrame(data)

    # 将DataFrame写入xlsx文件
    df.to_excel(file_path, index=False, engine='openpyxl')

    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    ws = wb.active
    set_column_widths(ws, df)
    # merge_cells(ws)
    merge_first_cells(ws)
    center_cells(ws)
    # add_checkboxes(ws, roles, df)

    # 保存工作簿
    wb.save(file_path)


# 合并重复的单元格
def merge_cells(ws):
    last_menu = None
    start_row = 2  # 从第二行开始，因为第一行是标题
    for row in range(2, ws.max_row + 1):
        current_menu = ws.cell(row=row, column=1).value
        if current_menu == last_menu:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row, end_column=1)
            # 设置合并单元格的对齐方式
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        else:
            last_menu = current_menu
            start_row = row


def merge_first_cells(ws):
    start_row = 1
    current_value = ws.cell(row=start_row, column=1).value

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value == current_value:
            continue
        else:
            # 如果起始行和当前行之间有间隔（即有重复值），则合并单元格
            if row - start_row > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
            # 更新起始行和当前值
            start_row = row
            current_value = cell.value

    # 处理最后一组重复值
    if ws.max_row - start_row > 0:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)


# 设置所有单元格居中对齐
def center_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


# 设置列宽度
def set_column_widths(ws, df):
    # 设置每列的宽度
    for idx, col in enumerate(df.columns):
        series = df[col]
        max_len = max((
            series.astype(str).map(len).max(),  # 获取列中字符串的最大长度
            len(str(series.name))  # 获取列名的长度
        )) + 1  # 增加一些额外的空间
        adjusted_width = (max_len + 2) * 1.2  # 调整宽度
        ws.column_dimensions[chr(65 + idx)].width = adjusted_width


# 添加复选框
def add_checkboxes(ws, roles, df):
    # 创建数据验证对象
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"')
    ws.add_data_validation(dv)

    # 添加复选框
    for role in roles:
        col_index = df.columns.get_loc(role.name) + 1  # 获取角色列的索引
        for row in range(2, ws.max_row + 1):  # 从第二行开始
            cell = ws.cell(row=row, column=col_index)
            dv.add(cell)
            if cell.value is True:
                cell.value = 'TRUE'
            elif cell.value is False:
                cell.value = 'FALSE'


if __name__ == '__main__':
    db_session = get_schema_session(org_code="mes_root")
    export_menus_to_excel(db_session, 'menus.xlsx')
